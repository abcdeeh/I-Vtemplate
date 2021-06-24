import openpyxl
from openpyxl.chart import Reference
from openpyxl.chart.axis import DateAxis
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
def add_to_chart(Sheet,file_name,p,Voltage,Current):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[Sheet]
    chart = ScatterChart()
    myfont = Font(typeface='Calibri')
    datawb = openpyxl.load_workbook(filename='data.xlsx')
    dataws = datawb.active

    cp = CharacterProperties(latin=myfont, sz=1400)
    chart.title = "I-V"
    chart.x_axis.title = 'V[V]'

    chart.y_axis.title = 'I[A]'

    for i in range(1,p+1):
        i_max_row=dataws.cell(row = 2, column = i ).value
        i_max_column=dataws.cell(row = 1, column = i ).value
        Accumulation=dataws.cell(row = 3, column = i  ).value
        frequency=dataws.cell(row = 4, column = i ).value

        X_axis = Reference(ws, min_col=Voltage+Accumulation, min_row=2, max_col=Voltage+Accumulation, max_row=i_max_row)
            #Y軸範囲
        Y_axis = Reference(ws, min_col=Accumulation+Current, min_row=2 , max_col=Accumulation+Current, max_row=i_max_row)

        con = Series(Y_axis, X_axis, title_from_data=False,title=frequency)
        con.marker.symbol = 'circle'
        con.marker.size = 4
        con.graphicalProperties.line.noFill = True
        chart.series.append(con)





    ws.add_chart(chart, "K7")
    #保存
    wb.save(file_name)
