import openpyxl
import numpy as np
def txt():
    datawb = openpyxl.load_workbook(filename='test.xlsx')
    dataws = datawb.active


    for j in range(1, dataws.max_row + 1):

        A=dataws.cell(row = j, column = 1 ).value


        B=A.split(' ')
        C=np.array(B)
        maxcolumn=C.shape[0]
        for i in range(1, maxcolumn + 1):
            H=C[i-1]
            if not H.isdecimal():
                if isfloat(H):
                    H=float(H)
            K=str(H)
            if K.isdecimal():
                H=int(H)

            dataws.cell(row = j, column = i ,value=H )
            datawb.save('test.xlsx')
def isfloat(string):
    try:
        float(string)
        return True
    except ValueError:
        return False
