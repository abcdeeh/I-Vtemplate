def paste(file_name,Sheet,frequency,source,delimiter_1,Accumulation,Current,p):
        import openpyxl
        from IV_programs import format
        import openpyxl
        import pandas as pd
        if delimiter_1==1:
            format.txt(source)
        if delimiter_1==3:
            format.csv(source)
        if delimiter_1==2:
            format.vba(source)

        excel_path_1="test.xlsx"
        #　コピーしたいファイル
        inwb = openpyxl.load_workbook(filename=excel_path_1)
        #　ペーストしたいファイル
        outwb = openpyxl.load_workbook(filename=file_name)
        datawb = openpyxl.load_workbook(filename='data.xlsx')
        # シートを選択（activeは現在開いているシート）
        outws=outwb[Sheet]
        inws = inwb.active
        dataws = datawb.active


        def copy_paste1():
            outws.cell(row = 1, column = Accumulation, value = frequency)
            for i in range(1, inws.max_row + 1):
                for j in range(1, inws.max_column + 1):
                    matrix = inws.cell(row = i, column = j).value
                    outws.cell(row = i, column = j+Accumulation, value = matrix)
                ba=outws.cell(row = i, column = Current+Accumulation).coordinate
                outws.cell(row = i, column = Accumulation+inws.max_column + 1, value ="="+ba+"/A4")
                i += 1
                outwb.save(file_name)
            outws.cell(row = 1, column = Accumulation+inws.max_column + 1, value = "J[A/cm^2]")

            outwb.save(file_name)
            dataws.cell(row = 1, column = p+1 , value = inws.max_column )
            dataws.cell(row = 2, column = p+1, value = inws.max_row )
            datawb.save('data.xlsx')



        copy_paste1()
